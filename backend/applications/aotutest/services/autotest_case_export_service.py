# -*- coding: utf-8 -*-
"""
@Author  : yangkai
@Email   : 807440781@qq.com
@Project : Krun
@Module  : autotest_case_export_service.py
@DateTime: 2026/7/27

测试用例导出服务：将「公共脚本」用例（单步骤 HTTP/TCP 请求）的请求头与请求体反向构造为
水平矩阵（HEAD/BODY 两段 + JSONPath 列），并生成 xlsx 工作簿。仅导出请求头与请求体，
不涉及断言（ASSERT_HEAD/ASSERT_BODY）。
"""
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from backend.common.convert_utils import Convert
from backend.configure import LOGGER
from backend.enums import AutoTestCaseType, AutoTestStepType, AutoTestReqArgsType

_SECTION_ORDER: Tuple[Tuple[str, str], ...] = (
    ("HEAD", "head_pairs"),
    ("BODY", "body_pairs"),
)

# 分栏标记集合（高亮用）：由 _SECTION_ORDER 派生，新增分段时自动纳入
_SECTION_MARKERS = {label for label, _ in _SECTION_ORDER}
# 分栏标记单元格高亮填充（主题橙）
_MARKER_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")


def build_export_file_name(username: Optional[str]) -> str:
    """
    构造导出文件名：{用户名}_接口数据_{时间戳}.xlsx（同步/异步导出共用，保证一致）。

    用户名做文件系统安全清洗（异步导出需落盘）；为空时省略用户名前缀。
    """
    safe_user = re.sub(r'[\\/:*?"<>|\s]', "_", str(username or "").strip())
    prefix = f"{safe_user}_接口数据" if safe_user else "接口数据"
    return f"{prefix}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"


def _kv_list_to_pairs(kv_list: Optional[List[Dict[str, Any]]]) -> List[Tuple[str, Any]]:
    """键值对列表([{key,value,desc}]) → [($.key, value)]，跳过空键。"""
    pairs: List[Tuple[str, Any]] = []
    for item in kv_list or []:
        key = item.get("key")
        if key is None or not str(key).strip():
            continue
        pairs.append((f"$.{str(key).strip()}", item.get("value")))
    return pairs


def flatten_to_jsonpath(data: Any, prefix: str = "$") -> List[Tuple[str, Any]]:
    """
    将嵌套 dict/list 扁平化为叶子 (JSONPath, 值) 列表。

    :param data: 待扁平化数据（dict/list）
    :param prefix: 当前路径前缀，根为 $
    :return: 形如 [("$.a.b", 1), ("$.c[0]", 2)] 的列表
    """
    pairs: List[Tuple[str, Any]] = []
    if isinstance(data, dict):
        for key, value in data.items():
            path = f"{prefix}.{key}"
            if isinstance(value, (dict, list)) and value:
                pairs.extend(flatten_to_jsonpath(value, path))
            else:
                pairs.append((path, "" if isinstance(value, (dict, list)) else value))
    elif isinstance(data, list):
        for index, value in enumerate(data):
            path = f"{prefix}[{index}]"
            if isinstance(value, (dict, list)) and value:
                pairs.extend(flatten_to_jsonpath(value, path))
            else:
                pairs.append((path, "" if isinstance(value, (dict, list)) else value))
    return pairs


def _body_to_pairs(step: Any) -> List[Tuple[str, Any]]:
    """按 request_args_type 取请求体并拆分为 JSONPath 键值对。"""
    args_type = getattr(step, "request_args_type", None)
    if args_type == AutoTestReqArgsType.JSON:
        body = getattr(step, "request_body", None)
        return flatten_to_jsonpath(body) if isinstance(body, dict) else []
    if args_type == AutoTestReqArgsType.XML:
        text = getattr(step, "request_text", None) or ""
        converted = Convert.xml_to_json(text) if text else ""
        if isinstance(converted, dict):
            return flatten_to_jsonpath(converted)
        return [("$.raw", text)] if text else []
    if args_type == AutoTestReqArgsType.FORM_DATA:
        return _kv_list_to_pairs(getattr(step, "request_form_data", None))
    if args_type == AutoTestReqArgsType.X_WWW_FORM_URLENCODED:
        return _kv_list_to_pairs(getattr(step, "request_form_urlencoded", None))
    if args_type == AutoTestReqArgsType.PARAMS:
        return _kv_list_to_pairs(getattr(step, "request_params", None))
    if args_type == AutoTestReqArgsType.RAW:
        text = getattr(step, "request_text", None) or ""
        return [("$.raw", text)] if text else []
    return []


def _collect_own_steps(steps: Optional[List[Any]]) -> List[Any]:
    """递归收集用例自身步骤（根步骤 + 子步骤，不含引用步骤）。"""
    collected: List[Any] = []
    for step in steps or []:
        collected.append(step)
        collected.extend(_collect_own_steps(getattr(step, "children", None)))
    return collected


async def prepare_export_cases(case_ids: List[int], services: Any) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    加载并校验待导出用例，拆分为请求矩阵键值对。

    校验规则（任一不满足即列入不合规清单）：用例存在、用例类型为公共脚本、用例步骤有且仅有 1 步、
    该步骤为 HTTP/TCP 请求步骤、步骤无数据源绑定。

    :param case_ids: 用例主键列表
    :param services: 自动化测试 CRUD 依赖聚合
    :return: (有效用例数据列表, 不合规清单[{case_id, case_name, reason}])
    """
    valid: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    project_name_cache: Dict[Any, str] = {}
    for case_id in case_ids:
        case = await services.case_curd.get_by_id(case_id)
        case_name = (getattr(case, "case_name", None) or str(case_id)) if case else str(case_id)
        if not case:
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "用例不存在"})
            continue
        if getattr(case, "case_type", None) != AutoTestCaseType.PUBLIC_SCRIPT:
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "非公共脚本用例"})
            continue
        load_result = await services.step_curd.get_by_case_id(case_id=case_id)
        own_steps = _collect_own_steps(getattr(load_result, "root_steps", None))
        if len(own_steps) != 1:
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": f"用例步骤数为{len(own_steps)}，需且仅需1步"})
            continue
        step = own_steps[0]
        step_type = getattr(step, "step_type", None)
        if step_type not in (AutoTestStepType.HTTP, AutoTestStepType.TCP):
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "步骤非HTTP/TCP请求步骤"})
            continue
        if getattr(step, "data_source_id", None):
            invalid.append({"case_id": case_id, "case_name": case_name, "reason": "步骤存在数据源绑定"})
            continue
        project_id = getattr(case, "case_project", None)
        if project_id not in project_name_cache:
            project = await services.project_curd.get_by_id(project_id)
            project_name_cache[project_id] = (getattr(project, "project_name", None) or "") if project else ""
        valid.append({
            "case_name": case_name,
            "protocol": "HTTP" if step_type == AutoTestStepType.HTTP else "TCP",
            "project_name": project_name_cache[project_id],
            "case_desc": getattr(case, "case_desc", None) or "",
            "created_user": getattr(case, "created_user", None) or "",
            "head_pairs": _kv_list_to_pairs(getattr(step, "request_header", None)),
            "body_pairs": _body_to_pairs(step),
        })
    LOGGER.info(f"导出用例准备完成: 有效{len(valid)}个, 不合规{len(invalid)}个")
    return valid, invalid


def build_case_matrix(case_data: Dict[str, Any]) -> List[List[Any]]:
    """
    构造单用例水平矩阵：第 0 行为分段标记+JSONPath 字段名，第 1 行为对应字段值。

    :param case_data: prepare_export_cases 产出的单用例数据
    :return: 二维矩阵
    """
    header: List[Any] = []
    data_row: List[Any] = []
    for label, field in _SECTION_ORDER:
        header.append(label)
        data_row.append("")
        for path, value in case_data.get(field) or []:
            header.append(path)
            data_row.append(value)
    return [header, data_row]


def _sanitize_sheet_name(name: Any, used: set) -> str:
    """清洗 Excel sheet 名（去非法字符、截断31、重名追加序号）。"""
    clean = re.sub(r"[:\\/?*\[\]]", "_", str(name or "").strip()) or "用例"
    clean = clean[:31]
    base = clean
    index = 1
    while clean in used:
        suffix = f"_{index}"
        clean = base[:31 - len(suffix)] + suffix
        index += 1
    used.add(clean)
    return clean


def _cell_safe(value: Any) -> Any:
    """单元格值规范化为 Excel 可写类型。"""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _style_sheet(sheet: Any) -> None:
    """统一设置 sheet 行高(30)/列宽(30)，并高亮 HEAD/BODY 等分栏标记所在单元格（无标记则跳过）。"""
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 30
        for cell in row:
            if cell.value in _SECTION_MARKERS:
                cell.fill = _MARKER_FILL
    for col_idx in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 30


def build_export_workbook(cases_data: List[Dict[str, Any]]) -> Workbook:
    """
    构建导出工作簿：每个用例一个数据 sheet；多于 1 个用例时末尾追加带超链接的「目录」sheet。

    :param cases_data: prepare_export_cases 产出的有效用例数据列表
    :return: openpyxl Workbook
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set = set()
    sheet_titles: List[Tuple[str, Dict[str, Any]]] = []
    for case_data in cases_data:
        title = _sanitize_sheet_name(case_data.get("case_name"), used_names)
        sheet = workbook.create_sheet(title=title)
        for row in build_case_matrix(case_data):
            sheet.append([_cell_safe(cell) for cell in row])
        sheet_titles.append((title, case_data))

    if len(sheet_titles) > 1:
        directory = workbook.create_sheet(title="目录", index=0)
        directory.append(["序号", "接口名称", "所属应用", "接口描述", "所属人", "协议类型"])
        for index, (title, case_data) in enumerate(sheet_titles, start=1):
            directory.append([
                index,
                case_data.get("case_name"),
                case_data.get("project_name"),
                case_data.get("case_desc"),
                case_data.get("created_user"),
                case_data.get("protocol"),
            ])
            # 超链接做在「接口名称」列，点击跳转对应数据 sheet
            name_cell = directory.cell(row=index + 1, column=2)
            name_cell.hyperlink = f"#'{title}'!A1"
            name_cell.style = "Hyperlink"
    for sheet in workbook.worksheets:
        _style_sheet(sheet)
    return workbook
