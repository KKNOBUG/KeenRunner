# -*- coding: utf-8 -*-
"""
@Author  : yangkai
@Email   : 807440781@qq.com
@Project : Krun
@Module  : autotest_xlsx_engine
@DateTime: 2026/4/10 15:31
"""
import asyncio
import re
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook

executor = ThreadPoolExecutor(max_workers=2)


def validate_excel_structure(sheets: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
    """
    校验多 sheet 数据驱动表结构：首行场景名一致、分区不重复、字段不重复。

    :param sheets: ``{sheet_name: DataFrame}``（header=None）
    :return: ``{"valid": bool, "error": list}``
    """
    errors = []
    first_row_standard = None

    for sheet_name, df in sheets.items():
        if df.empty:
            continue
        values = df.values

        current_first_row = tuple(values[0, 1:])
        if first_row_standard is None:
            first_row_standard = current_first_row
        else:
            if current_first_row != first_row_standard:
                errors.append(
                    {
                        "sheet": sheet_name,
                        "type": "首行场景名称顺序校验",
                        "section": "Header",
                        "row": 1,
                        "message": "第一行场景标题不一致"
                    }
                )
        first_col = values[:, 0]
        current_section = None
        section_fields = {
            "Head": set(),
            "Body": set(),
            "Assert-Head": set(),
            "Assert-Body": set()
        }
        section_start_row = None
        for i, cell in enumerate(first_col):
            excel_row = i + 1
            if not isinstance(cell, str):
                continue

            text_raw = cell.strip()
            if text_raw == "Head":
                if current_section == "Head":
                    errors.append(
                        {
                            "sheet": sheet_name,
                            "type": "接口请求信息",
                            "section": "Head",
                            "row": excel_row,
                            "message": "存在重复Head行"
                        }
                    )
                current_section = "Head"
                section_start_row = excel_row
                continue
            elif text_raw == "Body":
                if current_section == "Body":
                    errors.append(
                        {
                            "sheet": sheet_name,
                            "type": "接口请求信息",
                            "section": "Body",
                            "row": excel_row,
                            "message": "存在重复Body行"
                        }
                    )
                current_section = "Body"
                section_start_row = excel_row
                continue
            elif text_raw == "响应报文校验-Head":
                if current_section == "Assert-Head":
                    errors.append(
                        {
                            "sheet": sheet_name,
                            "type": "接口请求信息",
                            "section": "Assert",
                            "row": excel_row,
                            "message": "存在重复Assert行"
                        }
                    )
                current_section = "Assert-Head"
                section_start_row = excel_row
                continue
            elif text_raw == "响应报文校验-Body":
                if current_section == "Assert-Body":
                    errors.append(
                        {
                            "sheet": sheet_name,
                            "type": "接口请求信息",
                            "section": "Assert",
                            "row": excel_row,
                            "message": "存在重复Assert行"
                        }
                    )
                current_section = "Assert-Body"
                section_start_row = excel_row
                continue
            if current_section:
                field_key = text_raw
                if field_key in section_fields[current_section]:
                    errors.append(
                        {
                            "sheet": sheet_name,
                            "type": "接口请求信息",
                            "section": current_section,
                            "row": excel_row,
                            "message": f"{current_section}区域{text_raw}字段重复"
                        }
                    )
                else:
                    section_fields[current_section].add(field_key)
    return {
        "valid": len(errors) == 0,
        "error": errors
    }


def parse_kv_string(text: str, requests_body_key: Optional[str]) -> Dict[str, str]:
    """
    将多行 ``key: value`` 文本解析为 JSONPath 键值字典。

    :param text: 原始 KV 文本
    :param requests_body_key: 路径前缀片段；None 时前缀为 ``$.``
    :return: ``{JSONPath: value}``；非字符串输入返回空字典
    """
    if requests_body_key is None:
        key_path = "$."
    else:
        key_path = f"$.{requests_body_key}."
    if not isinstance(text, str):
        return {}
    result = {}
    lines = re.split(r"[\n\r]+", text)
    for line in lines:
        if ":" in line:
            k, v = line.split(":", 1)
            result[f"{key_path}{k.strip()}"] = v.strip()
    return result


def parse_sheet_fast(df: pd.DataFrame, sheet_name: str, requests_body_key: Dict[str, List[str]]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    同步解析单个 sheet 为根据场景组织的 head/body/assert 结构。

    :param df: sheet DataFrame（header=None）
    :param sheet_name: sheet 名，用于匹配 requests_body_key
    :param requests_body_key: sheet 到 body 路径键的映射
    :return: ``(场景字典, 错误列表)``
    """
    values = df.values

    scene_names = values[0, 1:]
    first_col = values[1:, 0]
    data_values = values[1:, 1:]

    sections = {"head": [], "body": [], "assert_head": [], "assert_body": []}
    sections_row_index = {"head": None, "body": None}
    current_section = None
    _label_map = {
        "head": "head",
        "body": "body",
        "assert_head": "assert_head",
        "assert_body": "assert_body",
    }

    for i, cell in enumerate(first_col):
        if not isinstance(cell, str):
            continue
        text = cell.strip().lower()
        section_key = _label_map.get(text)
        if section_key is not None:
            current_section = section_key
            if section_key in ("head", "body"):
                sections_row_index[section_key] = i
            continue
        if current_section:
            sections[current_section].append(i)

    result = {}
    errors = []
    col_count = data_values.shape[1]

    for col_idx in range(col_count):
        scene_name = scene_names[col_idx]
        if pd.isna(scene_name):
            continue

        record = {"head": {}, "body": {}, "assert_head": {}, "assert_body": {}}
        for section in ("head", "body"):
            row_idx = sections_row_index.get(section)
            if row_idx is None:
                continue
            raw_text = data_values[row_idx][col_idx]
            result_key = next((item for item in requests_body_key.get(sheet_name) if section in item.lower()), None)
            parsed_dict = parse_kv_string(raw_text, result_key)

            for k, v in parsed_dict.items():
                if k in record[section]:
                    errors.append(
                        {
                            "sheet": sheet_name,
                            "scene": scene_name,
                            "type": "接口请求字段",
                            "section": section,
                            "message": f"{section}区域{k}冲突"
                        }
                    )
                else:
                    record[section][f"{k.strip()}"] = v

        for section, rows in sections.items():
            for r in rows:
                key = first_col[r]
                value = data_values[r][col_idx]
                if key and not pd.isna(value):
                    if key in record[section]:
                        errors.append(
                            {
                                "sheet": sheet_name,
                                "scene": scene_name,
                                "type": "接口请求字段",
                                "section": section,
                                "row": r + 2,
                                "message": f"{section}区域{key}冲突"
                            }
                        )
                    else:
                        if section in ("head", "body"):
                            result_key = next((item for item in requests_body_key.get(sheet_name) if section in item.lower()), None)
                            if result_key is None:
                                key_path = "$."
                            else:
                                key_path = f"$.{result_key}."
                            record[section][f"{key_path}{key}"] = value
                        else:
                            record[section][key] = value
        result[scene_name] = record

    return result, errors


async def parse_sheet_async(df: pd.DataFrame, sheet_name: str, requests_body_key: Dict[str, List[str]]) -> Tuple[
    Dict[str, Any], List[Dict[str, Any]]]:
    """在线程池中异步调用 ``parse_sheet_fast``。"""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(executor, parse_sheet_fast, df, sheet_name, requests_body_key)


async def save_case_sheet(save_neo_name: Path, save_file_name: Path, sheet_name: str) -> None:
    """
    将源工作簿首个 sheet 复制到目标文件的指定 sheet（同名则覆盖）。

    :param save_neo_name: 目标 xlsx 路径
    :param save_file_name: 源 xlsx 路径
    :param sheet_name: 目标 sheet 名
    :return: None
    """
    source_wb = load_workbook(save_file_name)
    source_ws = source_wb[source_wb.sheetnames[0]]
    if save_neo_name.is_file():
        target_wb = load_workbook(save_neo_name)
    else:
        target_wb = Workbook()
        default_sheet = target_wb.active
        target_wb.remove(default_sheet)
    if sheet_name in target_wb.sheetnames:
        del target_wb[sheet_name]
    target_ws = target_wb.create_sheet(sheet_name)
    for row in source_ws.values:
        target_ws.append(row)
    target_wb.save(save_neo_name)


async def save_upload_file(upload_file: Any, destination: Path) -> None:
    """
    将上传文件分块写入目标路径。

    :param upload_file: 支持 ``read`` 的上传对象
    :param destination: 目标文件路径
    :return: None
    """
    with destination.open("wb") as buffer:
        while True:
            chunk = await upload_file.read(1024 * 1024)
            if not chunk:
                break
            buffer.write(chunk)
    return


async def xlsx_to_json_async(file_path: str, requests_body_key: Dict[str, List[str]], first_sheet_only: bool = False) -> Dict[str, Any]:
    """
    异步读取 xlsx，校验结构后解析为根据 sheet/场景组织的 JSON。

    :param file_path: Excel 路径
    :param requests_body_key: sheet 到 body 路径键的映射
    :param first_sheet_only: 为 True 时仅解析首个 sheet
    :return: ``{"valid": True, "data": ...}`` 或 ``{"valid": False, "errors": ...}``
    """
    if first_sheet_only:
        df = pd.read_excel(file_path, sheet_name=0, header=None)
        sheets = {"sheet1": df}
    else:
        sheets = pd.read_excel(file_path, sheet_name=None, header=None)
    validate_result = validate_excel_structure(sheets)
    if not validate_result["valid"]:
        return validate_result
    tasks = [
        parse_sheet_async(df, sheet_name, requests_body_key)
        for sheet_name, df in sheets.items()
        if not df.empty
    ]

    # for sheet_name, df in sheets.items():
    #     tasks.append(parse_sheet_async(df))

    results = await asyncio.gather(*tasks)
    final_data = {}
    all_error = []
    for (sheet_name, _), (data, errors) in zip(sheets.items(), results):
        final_data[sheet_name] = data
        all_error.extend(errors)
    if all_error:
        return {
            "valid": False,
            "errors": all_error
        }
    return {
        "valid": True,
        "data": final_data
    }
